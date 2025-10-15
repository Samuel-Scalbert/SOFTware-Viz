import os
import json
from tqdm import tqdm
import xml.etree.ElementTree as ET
import requests
from openpyxl import load_workbook
from Utils.TEI_to_JSON import transformer_TEI_JSON
from Utils.elastic_search import sync_to_elasticsearch

def check_or_create_collection(db, collection_name, collection_type='Collection'):
    """
    Checks if a collection exists in the database. If not, creates the collection.
    :param db: Database connection
    :param collection_name: Name of the collection
    :param collection_type: Type of the collection ('Collection' or 'Edges')
    :return: The collection object
    """
    if db.hasCollection(collection_name):
        return db[collection_name]
    else:
        db.createCollection(collection_type, name=collection_name)
        return db[collection_name]

def duplicates_JSON(lst):
    seen = set()
    duplicates = []

    for item in lst:
        item_hashable = str(item)
        if item_hashable in seen:
            duplicates.append(item)
        else:
            seen.add(item_hashable)

    return duplicates


def find_ancestor_paths(current_affiliation_id, ns, tree):
    # Look for the current structure's relations
    list_relation = tree.find(
        f".//tei:back//tei:listOrg//tei:org[@xml:id='{current_affiliation_id}']//tei:listRelation", ns)

    # If no relations, we're at the top, return the current structure as the only member of the path
    if list_relation is None:
        return [[
            current_affiliation_id]]  # Return a list with a single path containing the current structure

    # Collect all direct parent relations (and their types)
    direct_parents = []
    for relation in list_relation:
        if relation.attrib.get('type') == 'direct':
            parent_id = relation.attrib['active'][1:]  # Strip the initial "#" for the parent ID
            direct_parents.append(parent_id)

    # If no direct parents, return the current structure as the topmost node
    if not direct_parents:
        return [[current_affiliation_id]]  # Return a path with only the current structure

    # For each direct parent, find its ancestor paths and create a separate list for each path
    all_paths = []
    for parent_id in direct_parents:
        parent_paths = find_ancestor_paths(parent_id, ns, tree)  # Recursive call for each parent
        for path in parent_paths:
            all_paths.append(
                [current_affiliation_id] + path)  # Create a separate path for each relation
    return all_paths  # Return all paths for the current structure

def insert_json_db(data_path_json,data_path_xml,db):
    software_document = []
    list_errors = []
    #db.dropAllCollections()

    workbook = load_workbook(filename='./app/static/data/Logiciels_Blacklist_et_autres_remarques.xlsx')
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    blacklist = []
    for row in data[1:]:
        blacklist.append(row[0])

    #load the urls verified with SH
    try:
        with open('app/static/data/url_verified_with_SH/urls.json', 'r') as file:
            urls_verified_sh = json.load(file)
    except FileNotFoundError:
        urls_verified_sh = []

    # Create or retrieve collections
    documents_collection = check_or_create_collection(db, 'documents')
    softwares_collection = check_or_create_collection(db, 'softwares')
    references_collection = check_or_create_collection(db, 'references')
    authors_collection = check_or_create_collection(db, 'authors')
    structures_collection = check_or_create_collection(db, 'structures')
    list_relation_collection = check_or_create_collection(db, 'list_relation')

    # Create or retrieve edge collections
    doc_struc_edge = check_or_create_collection(db, 'edge_doc_to_struc', 'Edges')
    auth_struc_edge = check_or_create_collection(db, 'edge_auth_to_struc', 'Edges')
    doc_soft_edge = check_or_create_collection(db, 'edge_doc_to_software', 'Edges')
    doc_ref_edge = check_or_create_collection(db, 'edge_doc_to_reference', 'Edges')
    doc_auth_edge = check_or_create_collection(db, 'edge_doc_to_author', 'Edges')
    auth_rel_struc_edge = check_or_create_collection(db, 'edge_auth_to_rel_struc', 'Edges')

    data_json_files = os.listdir(data_path_json)
    data_xml_list = os.listdir(data_path_xml)
    files_list_registered = db.AQLQuery('FOR hal_id in documents RETURN hal_id.file_hal_id', rawResults=True, batchSize=2000)
    dict_registered = {}
    dict_edge_author = {}
    new_file = False

    for data_file_xml in tqdm(data_xml_list[:0]):
        file_path = f'{data_path_xml}/{data_file_xml}'
        file_name = os.path.basename(file_path)
        while "." in file_name:
            file_name, extension = os.path.splitext(file_name)



        if file_name in files_list_registered:
            continue
        else:
            new_file = True

        url = "https://api.archives-ouvertes.fr/search/"
        params = {
            "q": f"halId_id:{file_name}",
            "rows": 10,
            "fl": "citationFull_s"
        }

        response = requests.get(url, params=params)
        if response.status_code == 200:
            data = response.json()
            try:
                citations = data["response"]["docs"][0]["citationFull_s"]
            except IndexError:
                citations = None
        else:
            print("Error:", response.status_code)
        with open(file_path, 'r', encoding='utf-8') as xml_file:
            data_json_get_document = {}
            tree = ET.parse(xml_file)
            root = tree.getroot()
            ns = {"tei": "http://www.tei-c.org/ns/1.0", 'xml': 'http://www.w3.org/XML/1998/namespace'}


            # DOCUMENT -----------------------------------------------------

            if file_name in urls_verified_sh:
                urls_list = urls_verified_sh[file_name]
                data_json_get_document['urls_verified_SH'] = urls_list

            title = tree.find(".//tei:listBibl//tei:titleStmt//tei:title", ns)
            title = title.text

            doc_type = tree.findall(".//tei:listBibl//tei:biblFull//tei:profileDesc//tei:textClass//tei:classCode", ns)
            for tag in doc_type:
                if tag.attrib.get('n') == 'COMM':
                    production_date = tree.find(".//tei:listBibl//tei:biblFull//tei:sourceDesc//tei:biblStruct//tei:monogr//tei:meeting//tei:date[@type='start']", ns)
                    if production_date is not None:
                        data_json_get_document['date'] = production_date.text[:4]
                else:
                    production_date = tree.find(
                        ".//tei:listBibl//tei:biblFull//tei:sourceDesc//tei:biblStruct//tei:monogr//tei:imprint//tei:date[@type='datePub']",
                        ns)
                    if production_date is not None:
                        data_json_get_document['date'] = production_date.text[:4]
                    else:
                        production_date = tree.find(
                            ".//tei:listBibl//tei:biblFull//tei:editionStmt//tei:edition[@type='current']//tei:date[@type='whenProduced']",
                            ns)
                        if production_date is not None:
                            data_json_get_document['date'] = production_date.text[:4]
                        else:
                            print(f'problem : {file_name}')


            abstract = tree.find(".//{http://www.tei-c.org/ns/1.0}abstract")
            if abstract:
                tag_text = list(abstract)[0]
                if tag_text.tag == '{http://www.tei-c.org/ns/1.0}p':
                    abstract = "".join(tag_text.itertext())
                    data_json_get_document['abstract'] = ['HAL' , abstract]
                if tag_text.tag == '{http://www.tei-c.org/ns/1.0}div':
                    for p_tag in list(tag_text):
                        text = "".join(p_tag.itertext())
                    data_json_get_document['abstract'] = ['GROBID' , abstract]

            data_json_get_document['file_hal_id'] = file_name
            data_json_get_document['citation'] = citations
            data_json_get_document['title'] = title

            document_document = documents_collection.createDocument(data_json_get_document)
            document_document.save()

            # SOFTWARE -----------------------------------------------------

            if f"{file_name}.software.json" in data_json_files:
                with open(f'{data_path_json}/{file_name}.software.json', 'r') as json_file:
                    data_json = json.load(json_file)
                    data_json_get_mentions = data_json.get("mentions")

                    # Remove duplicates
                    for elm in duplicates_JSON(data_json_get_mentions):
                        data_json_get_mentions.remove(elm)

                    # Process each mention
                    for mention in data_json_get_mentions:
                        if mention['software-name']['normalizedForm'] not in blacklist:
                            mention['software_name'] = mention.pop('software-name')
                            mention['software_type'] = mention.pop('software-type')
                            software_document = softwares_collection.createDocument(mention)
                            software_document.save()

                            # Create edge from document to software
                            edge_doc_soft = doc_soft_edge.createEdge()
                            edge_doc_soft['_from'] = document_document._id
                            edge_doc_soft['_to'] = software_document._id
                            edge_doc_soft.save()

            # REFERENCES -----------------------------------------------------

                    # Process each reference
                    data_json_get_references = data_json.get("references")
                    for reference in data_json_get_references:
                        result_json = []
                        try:
                            result_json,error = transformer_TEI_JSON(reference['tei'])
                            if len(error) > 0:
                                list_errors.append(error, reference['tei'])
                        except Exception as e:
                            print(f"Error during the transformation from XML to JSON: {e}")
                        if result_json:
                            reference['json'] = result_json
                        references_document = references_collection.createDocument(reference)
                        references_document.save()

                        # Create edge from document to reference
                        edge_doc_ref = doc_ref_edge.createEdge()
                        edge_doc_ref['_from'] = document_document._id
                        edge_doc_ref['_to'] = references_document._id
                        edge_doc_ref.save()
                # Define the AQL query to fetch software names and their counts
                query = f"""
                FOR doc IN edge_software
                    FILTER doc._from == "{document_document._id}"
                    LET software = DOCUMENT(doc._to)
                    COLLECT softwareName = software.software_name.normalizedForm WITH COUNT INTO count
                    RETURN {{ softwareName, count }}
                """

            # Cleaning the software for the dashboard (maybe obsolete now) -----------------------------------------

                '''# Execute the AQL query to get software names and counts
                all_software_dict = db.AQLQuery(query, rawResults=True)

                # Fetch distinct software names
                distinct_query = f"""
                FOR doc IN edge_software
                    FILTER doc._from == "{document_document._id}"
                    LET software = DOCUMENT(doc._to)
                    RETURN DISTINCT software.software_name.normalizedForm
                """
                software_name_list = db.AQLQuery(distinct_query, rawResults=True)

                # Convert software names and counts to a dictionary
                dict_software = {software_dict['softwareName']: software_dict['count'] for software_dict in
                                 all_software_dict}

                # Process software names containing hyphens and update as needed
                for software_name in software_name_list:
                    if "-" in software_name:
                        software_name_cleaned = software_name.replace('-', '')
                        if software_name_cleaned in dict_software and dict_software[software_name] < dict_software[
                            software_name_cleaned]:
                            # Example: Update software name in the database
                            update_query = f"""
                            FOR doc IN edge_software
                                FILTER doc._from == "{document_document._id}"
                                LET software = DOCUMENT(doc._to)
                                FILTER software.software_name.normalizedForm == "{software_name}"
                                UPDATE software WITH {{ software_name: {{ normalizedForm: "{software_name_cleaned}" }} }} IN softwares
                            """
                            db.AQLQuery(update_query)'''

            # AUTHORS -----------------------------------------------------

            author_list = tree.findall(".//tei:listBibl//tei:titleStmt//tei:author", ns)
            list_author_old = []
            for elm in author_list:
                author = {}
                persName = elm.find("{http://www.tei-c.org/ns/1.0}persName")
                author['role'] = elm.attrib['role']
                author = {}

                # idhal
                author_id = {}
                id_halid = None
                id_halauthorid = None
                id_halid = elm.findall(".//tei:idno[@type='idhal']", ns)
                if len(id_halid) > 0:
                    for id_type in id_halid:
                        author_id[id_type.attrib['notation']] = id_type.text
                id_halauthorid = elm.find(".//tei:idno[@type='halauthorid']", ns)
                if id_halauthorid is not None:
                    author_id[id_halauthorid.attrib['type']] = id_halauthorid.text
                    id_final = 0
                try:
                    author_final_id = author_id['numeric']
                except KeyError:
                    author_final_id = author_id['halauthorid']
                if author_final_id in list(dict_registered.keys()):
                    registered = True
                else:
                    registered = False

                # name
                author_name = {}
                persName = elm.find("{http://www.tei-c.org/ns/1.0}persName")
                for name in persName:
                    author_name[name.tag.split('}')[1]] = name.text

                # document
                author_documents = []
                try:
                    # Create a dictionary for each document and append it to the list
                    document_data = {
                        'document_halid': data_file_xml.replace(".hal.xml", "").replace(".hal.grobid.xml", "").replace(".xml", "").replace(".hal.xml", ""),
                        'role': elm.attrib['role']
                    }
                    author_documents.append(document_data)
                except KeyError:
                    # Handle the case where 'role' may not be present, and append with a default 'role'
                    document_data = {
                        'document_halid': data_file_xml.replace(".hal.xml", "").replace(".hal.grobid.xml", "").replace(".xml", "").replace(".hal.xml", ""),
                        'role': elm.attrib.get('role', 'unknown')
                    }
                    author_documents.append(document_data)

                if registered == False:

                    author['id'] = author_id
                    author['name'] = author_name
                    author['documents'] = author_documents

                    document_author = authors_collection.createDocument(author)
                    document_author.save()
                    dict_registered[author_final_id] = document_author._id
                    # print(dict_registered)
                    edge_doc_auth = doc_auth_edge.createEdge()
                    edge_doc_auth['_from'] = document_document._id
                    edge_doc_auth['_to'] = document_author._id
                    edge_doc_auth.save()
                    author_document_id = document_author._id

                elif registered == True:
                    author_document_id = dict_registered[author_final_id]
                    # AQL query to append to the documents and affiliation
                    aql_query = f'''
                                FOR doc IN authors
                                    FILTER doc._id == '{author_document_id}'
                                    UPDATE doc WITH {{ 
                                        documents: APPEND(doc.documents, {author_documents}, true), 
                                    }} IN authors
                                '''
                    # Execute the AQL query
                    result = db.AQLQuery(aql_query, rawResults=True)
                    edge_doc_auth = doc_auth_edge.createEdge()
                    edge_doc_auth['_from'] = document_document._id
                    edge_doc_auth['_to'] = author_document_id
                    edge_doc_auth.save()
            # STRUCT -----------------------------------------------------

                # Main logic to collect paths for all affiliations
                author_affiliation_paths = []
                affiliations_list = elm.findall("{http://www.tei-c.org/ns/1.0}affiliation")
                if len(affiliations_list) > 0:
                    for affiliation in affiliations_list:
                        # Start with the affiliation's referenced structure
                        current_affiliation_id = affiliation.attrib['ref'][1:]  # Strip the initial "#"
                        ancestor_paths = find_ancestor_paths(current_affiliation_id, ns,
                                                             tree)  # Get all full hierarchy paths
                        author_affiliation_paths.extend(ancestor_paths)  # Add all paths to the final list
                        # Dictionary to store document key to _id mappings
                        document_id_map = {}
                        for author_affiliation_path in author_affiliation_paths:
                            list_relation_idstruct = []
                            for index, structure in enumerate(author_affiliation_path):
                                result = db.AQLQuery(
                                    f'FOR struc IN structures FILTER struc.id_haureal == "{structure}" RETURN struc._id',
                                    rawResults=True)
                                if len(result) == 0:
                                    affiliated_struct = tree.find(
                                        f".//tei:back//tei:listOrg//tei:org[@xml:id='{structure}']", ns)
                                    if affiliated_struct is not None:
                                        affiliated_struct_name = affiliated_struct.findall('tei:orgName', ns)
                                        if affiliated_struct_name is not None:
                                            org = {}
                                            org['status'] = affiliated_struct.attrib['status']
                                            org['url_team'] = False
                                            url_team = affiliated_struct.find(
                                                './/{http://www.tei-c.org/ns/1.0}desc//ref[@type="url"]')
                                            for url_team in affiliated_struct.find(
                                                    './/{http://www.tei-c.org/ns/1.0}desc'):
                                                if url_team.attrib == {'type': 'url'}:
                                                    org['url_team'] = url_team.text
                                            try:
                                                list_affiliated_str = list(affiliated_struct.find(
                                                    ".//{http://www.tei-c.org/ns/1.0}listRelation"))
                                            except TypeError:
                                                list_affiliated_str = []
                                            for struct in affiliated_struct_name:
                                                if struct.attrib:
                                                    org[struct.attrib['type']] = struct.text
                                                else:
                                                    org["name"] = struct.text
                                            org['id_haureal'] = affiliated_struct.attrib["{http://www.w3.org/XML/1998/namespace}id"]
                                            org['type'] = affiliated_struct.attrib['type']
                                            document_org = structures_collection.createDocument(org)
                                            document_org.save()
                                            struct_id = document_org._id
                                else:
                                    struct_id = result[0]
                                    pass

                                query = f"""
                                            for edge in edge_doc_to_struc
                                                filter edge._from == "{document_document._id}"
                                                let struct = document(edge._to)
                                                return struct._id
                                                """
                                list_registered_edge_struc_doc = db.AQLQuery(query, rawResults=True)

                                if struct_id not in list_registered_edge_struc_doc:
                                    edge_doc_struc = doc_struc_edge.createEdge()
                                    edge_doc_struc['_from'] = document_document._id
                                    edge_doc_struc['_to'] = struct_id
                                    edge_doc_struc.save()

                                query = f"""
                                            for edge in edge_auth_to_struc
                                                filter edge._from == "{author_document_id}"
                                                let struct = document(edge._to)
                                                return struct._id
                                                """
                                list_registered_edge_struc_author = db.AQLQuery(query, rawResults=True)

                                if struct_id not in list_registered_edge_struc_author:
                                    edge_auth_struc = auth_struc_edge.createEdge()
                                    edge_auth_struc['_from'] = author_document_id
                                    edge_auth_struc['_to'] = struct_id
                                    edge_auth_struc.save()

                                result = db.AQLQuery(
                                    f'FOR struc IN structures FILTER struc.id_haureal == "{structure}" RETURN struc._id',
                                    rawResults=True)
                                if result != []:
                                    list_relation_idstruct.append(result[0])

                            list_relation_documents = list_relation_collection.createDocument({"list_relation" : list_relation_idstruct})
                            list_relation_documents.save()

                        edge_auth_rel_struc = auth_rel_struc_edge.createEdge()
                        edge_auth_rel_struc['_from'] = author_document_id
                        edge_auth_rel_struc['_to'] = list_relation_documents._id
                        edge_auth_rel_struc.save()
    if new_file == True:
        sync_to_elasticsearch(db)
    if len(list_errors) > 0:
       print(list_errors)
